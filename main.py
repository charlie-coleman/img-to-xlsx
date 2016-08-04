from PIL import Image
import openpyxl as xl
import tkinter as tk
from tkinter import filedialog
from openpyxl.styles import PatternFill
import time


tk.Tk().withdraw()
filename = tk.filedialog.askopenfilename(initialdir='.\\pictures\\')
suggested_name = filename[filename.rindex(r'/')+1:filename.rindex(r'.')]
file_output = tk.filedialog.asksaveasfilename(defaultextension='.xlsx',
                                              initialfile=suggested_name,
                                              initialdir='.\\excelfiles\\')
start = time.time()


def rgb_to_hex(rgb):
    return 'ff%02x%02x%02x' % rgb

image = Image.open(filename)
wb = xl.Workbook()
ws = wb.active
greater = image.height #if image.height > image.width else image.width
max_size = 782
if greater > max_size:
    scale_factor = max_size/greater
else:
    scale_factor = 1
image = image.resize([int(scale_factor * s) for s in image.size])
for x in range(0, image.width):
    for y in range(0, image.height):
        color = rgb_to_hex(image.getpixel((x, y))[:3])
        ws.cell(row=y+1, column=x+1).fill = PatternFill(
            start_color=color,
            end_color=color,
            fill_type='solid'
        )
print(time.time()-start)
wb.save(file_output)
print(time.time()-start)
